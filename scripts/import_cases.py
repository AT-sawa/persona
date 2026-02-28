#!/usr/bin/env python3
"""Import cases from Excel '案件 のコピー' sheet into Supabase."""

import openpyxl
import re
import json
import sys
from datetime import datetime
from urllib.request import Request, urlopen

SUPABASE_URL = "https://urikwrakbafnsllimcbl.supabase.co"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVyaWt3cmFrYmFmbnNsbGltY2JsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MjA4ODY1NSwiZXhwIjoyMDg3NjY0NjU1fQ.GVfOSM9ZBpWdOchnLb8PYDLZkrMU6LCA1i1Z4mXivUo"

EXCEL_FILE = "フリーコンサル案件登録.xlsx"
SHEET_NAME = "案件 のコピー"

# Statuses to skip (duplicates, issues)
SKIP_KEYWORDS = ["同案件", "同内容", "スキップ", "入力できず", "登録せず", "対象外", "非掲載"]


def parse_bracket_format(content: str) -> dict:
    """Parse [field]value or 【field】value format."""
    result = {}
    # Half-width brackets
    patterns = {
        "title": r"\[案件名\](.+?)(?:\n|\[)",
        "flow": r"\[商流\](.+?)(?:\n|\[)",
        "fee": r"\[単価[^\]]*\](.+?)(?:\n|\[)",
        "location": r"\[作業場所\](.+?)(?:\n|\[)",
        "start_date": r"\[契約開始日\](.+?)(?:\n|\[)",
        "extendable": r"\[継続可能性\](.+?)(?:\n|\[)",
        "occupancy": r"\[稼働率[^\]]*\](.+?)(?:\n|\[)",
        "description": r"\[作業内容\]\s*\n?([\s\S]+?)(?:\[必須スキル\]|\[尚可スキル\]|\[補足事項\]|$)",
        "must_req": r"\[必須スキル\]\s*\n?([\s\S]+?)(?:\[尚可スキル\]|\[補足事項\]|$)",
        "nice_to_have": r"\[尚可スキル\]\s*\n?([\s\S]+?)(?:\[補足事項\]|$)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, content)
        if m:
            result[key] = m.group(1).strip()
    return result


def parse_fullwidth_bracket_format(content: str) -> dict:
    """Parse 【field】value format (newer cases)."""
    result = {}
    patterns = {
        "title": r"【案件名】\s*(.+?)(?:\n|【)",
        "flow": r"【商流】\s*(.+?)(?:\n|【)",
        "fee": r"【単価】\s*(.+?)(?:\n|【)",
        "location": r"【メイン勤務地】\s*(.+?)(?:\n|【)",
        "start_date": r"【案件期間】\s*(.+?)(?:\n|【)",
        "occupancy": r"【稼働率】\s*(.+?)(?:\n|【)",
        "office_days": r"【出社頻度】\s*(.+?)(?:\n|【)",
        "must_req": r"【必須スキル】\s*\n?([\s\S]+?)(?=【尚可スキル】|【募集人数】|【出社頻度】|【年齢】|【英語|$)",
        "nice_to_have": r"【尚可スキル】\s*\n?([\s\S]+?)(?=【募集人数】|【出社頻度】|【年齢】|【英語|$)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, content)
        if m:
            result[key] = m.group(1).strip()

    # Description from 【募集要項】 or 【案件概要】
    for tag in ["【募集要項】", "【案件概要】"]:
        m = re.search(re.escape(tag) + r"\s*\n?([\s\S]+?)(?=【単価】|【商流】|【必須スキル】|$)", content)
        if m:
            result["description"] = m.group(1).strip()
            break

    # Also handle ■ format mixed in
    if not result.get("description"):
        m = re.search(r"■概要\s*\n?([\s\S]+?)(?=■必須|■条件|■尚可|【|$)", content)
        if m:
            result["description"] = m.group(1).strip()
    if not result.get("must_req"):
        m = re.search(r"■必須\s*\n?([\s\S]+?)(?=■尚可|■条件|【|$)", content)
        if m:
            result["must_req"] = m.group(1).strip()

    # Extract fee from ■条件面 section if not found
    if not result.get("fee"):
        m = re.search(r"【単価】\s*(.+?)(?:\n|$)", content)
        if m:
            result["fee"] = m.group(1).strip()

    return result


def parse_blacksquare_format(content: str) -> dict:
    """Parse ■field：value format."""
    result = {}
    patterns = {
        "title": r"■案件名[：:](.+?)(?:\n|■|$)",
        "description": r"■案件概要[：:]\s*\n?([\s\S]+?)(?=■必須|■ポジション|■尚可|■開始|■稼働|■単価|■勤務|■年齢|$)",
        "must_req": r"■必須スキル[：:]\s*\n?([\s\S]+?)(?=■尚可|■開始|■稼働|■単価|■勤務|■年齢|$)",
        "nice_to_have": r"■尚可スキル[：:]\s*\n?([\s\S]+?)(?=■開始|■稼働|■単価|■勤務|■年齢|$)",
        "fee": r"■単価[：:](.+?)(?:\n|■|$)",
        "location": r"■勤務場所[：:](.+?)(?:\n|■|$)",
        "occupancy": r"■稼働率[：:](.+?)(?:\n|■|$)",
        "start_date": r"■開始日[：:](.+?)(?:\n|■|$)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, content)
        if m:
            result[key] = m.group(1).strip()
    return result


def parse_diamond_format(content: str) -> dict:
    """Parse ◆field：value format."""
    result = {}
    patterns = {
        "title": r"◆概要[：:](.+?)(?:\n|◆|$)",
        "flow": r"◆元請け[：:](.+?)(?:\n|◆|$)",
        "fee": r"◆単価[：:](.+?)(?:\n|◆|$)",
        "location": r"◆勤務地[：:](.+?)(?:\n|◆|$)",
        "start_date": r"◆期間[：:](.+?)(?:\n|◆|$)",
        "extendable": r"◆延長予定[：:](.+?)(?:\n|◆|$)",
        "occupancy": r"◆稼働率[：:](.+?)(?:\n|◆|$)",
        "description": r"◆内容[：:]\s*\n?([\s\S]+?)(?=◆必須スキル|◆尚可スキル|◆期間|◆勤務地|$)",
        "must_req": r"◆必須スキル[：:]\s*\n?([\s\S]+?)(?=◆尚可スキル|◆期間|◆勤務地|◆稼働率|◆単価|$)",
        "nice_to_have": r"◆尚可スキル[：:]\s*\n?([\s\S]+?)(?=◆期間|◆勤務地|◆稼働率|◆単価|$)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, content)
        if m:
            result[key] = m.group(1).strip()

    # Industry from エンドクライアント
    m = re.search(r"◆エンドクライアント[：:](.+?)(?:\n|◆|$)", content)
    if m:
        result["industry"] = m.group(1).strip()

    return result


def parse_content(content: str) -> dict:
    """Parse case content from either format."""
    if not content:
        return {}
    content = str(content)
    if "【案件名】" in content or "【単価】" in content or "【募集要項】" in content:
        return parse_fullwidth_bracket_format(content)
    elif "■案件名" in content or ("■案件概要" in content and "■単価" in content):
        return parse_blacksquare_format(content)
    elif "■概要" in content or "■必須" in content:
        return parse_fullwidth_bracket_format(content)
    elif "[案件名]" in content:
        return parse_bracket_format(content)
    elif "◆" in content:
        return parse_diamond_format(content)
    else:
        result = parse_bracket_format(content)
        if not result.get("title") and not result.get("fee"):
            result = parse_diamond_format(content)
        if not result.get("title") and not result.get("fee"):
            result = parse_fullwidth_bracket_format(content)
        if not result.get("title") and not result.get("fee"):
            result = parse_blacksquare_format(content)
        return result


def determine_active(req_date, close_date) -> bool:
    """Determine if a case should be active."""
    if close_date:
        return False
    if not isinstance(req_date, datetime):
        return False
    return req_date.year >= 2025


def determine_category(excel_cat, content: str) -> str:
    """Determine case category."""
    if excel_cat:
        cat = str(excel_cat).strip()
        if cat in ("コンサル", "SI"):
            return cat
    return "コンサル"


def extract_fee(fee_str: str) -> str:
    """Normalize fee string."""
    if not fee_str:
        return ""
    fee_str = fee_str.strip()
    if "万" in fee_str:
        return fee_str
    nums = re.findall(r"[\d,.]+", fee_str)
    if nums:
        return f"{'-'.join(nums)}万円/月"
    return fee_str


def get_existing_case_titles():
    """Fetch all existing case titles from DB."""
    url = f"{SUPABASE_URL}/rest/v1/cases?select=title,case_no"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    }
    req = Request(url, headers=headers)
    resp = urlopen(req)
    data = json.loads(resp.read().decode())
    titles = set()
    case_nos = set()
    for row in data:
        if row.get("title"):
            titles.add(row["title"].strip())
        if row.get("case_no"):
            case_nos.add(str(row["case_no"]).strip())
    return titles, case_nos


def main():
    print("Loading Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    print("Fetching existing cases from DB...")
    existing_titles, existing_nos = get_existing_case_titles()
    print(f"Existing: {len(existing_titles)} titles, {len(existing_nos)} case_nos")

    cases_to_import = []
    skipped_dup = 0
    skipped_status = 0
    skipped_no_title = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        content = row[5]  # 登録内容
        if not content:
            continue

        status = str(row[7]) if row[7] else ""
        if any(kw in status for kw in SKIP_KEYWORDS):
            skipped_status += 1
            continue

        req_date = row[1]  # 本件ご依頼日
        close_date = row[8]  # クローズ日
        excel_cat = row[2]  # 案件カテゴリ
        row_no = row[0]  # No.
        col_title = row[4]  # 案件タイトル（タイトルのみ）

        parsed = parse_content(str(content))

        # Determine title
        title = ""
        if col_title:
            title = str(col_title).strip()
        elif parsed.get("title"):
            title = parsed["title"]
        else:
            first_line = str(content).split("\n")[0].strip()
            for prefix in ["[案件名]", "◆概要：", "◆概要:"]:
                first_line = first_line.replace(prefix, "")
            title = first_line[:200]

        if not title:
            skipped_no_title += 1
            continue

        # Check duplicates
        case_no = str(row_no).strip() if row_no else ""
        if title.strip() in existing_titles:
            skipped_dup += 1
            continue

        is_active = determine_active(req_date, close_date)
        category = determine_category(excel_cat, str(content))
        fee = extract_fee(parsed.get("fee", ""))

        pub_date = None
        if isinstance(req_date, datetime):
            pub_date = req_date.isoformat()

        case = {
            "case_no": case_no if case_no else None,
            "title": title[:500],
            "category": category,
            "description": (parsed.get("description") or "")[:5000] or None,
            "fee": fee or None,
            "location": (parsed.get("location") or "")[:500] or None,
            "occupancy": (parsed.get("occupancy") or "")[:200] or None,
            "must_req": (parsed.get("must_req") or "")[:5000] or None,
            "nice_to_have": (parsed.get("nice_to_have") or "")[:5000] or None,
            "flow": (parsed.get("flow") or "")[:200] or None,
            "start_date": (parsed.get("start_date") or "")[:200] or None,
            "extendable": (parsed.get("extendable") or "")[:200] or None,
            "industry": (parsed.get("industry") or "")[:200] or None,
            "is_active": is_active,
            "status": "active" if is_active else "closed",
            "published_at": pub_date,
        }

        existing_titles.add(title.strip())
        cases_to_import.append(case)

    print(f"\nParsing complete:")
    print(f"  To import: {len(cases_to_import)}")
    print(f"  Skipped (status): {skipped_status}")
    print(f"  Skipped (duplicate): {skipped_dup}")
    print(f"  Skipped (no title): {skipped_no_title}")
    print(f"  Active: {sum(1 for c in cases_to_import if c['is_active'])}")
    print(f"  Inactive: {sum(1 for c in cases_to_import if not c['is_active'])}")

    if "--dry-run" in sys.argv:
        with open("cases_preview.json", "w") as f:
            json.dump(cases_to_import[:20], f, ensure_ascii=False, indent=2)
        print("\nDry run - saved first 20 to cases_preview.json")
        return

    # Insert in batches of 100
    print(f"\nInserting {len(cases_to_import)} cases...")
    batch_size = 100
    inserted = 0
    errors = 0

    for i in range(0, len(cases_to_import), batch_size):
        batch = cases_to_import[i : i + batch_size]
        url = f"{SUPABASE_URL}/rest/v1/cases"
        headers = {
            "apikey": SUPABASE_SERVICE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }
        body = json.dumps(batch, ensure_ascii=False).encode("utf-8")
        req = Request(url, data=body, headers=headers, method="POST")
        try:
            resp = urlopen(req)
            inserted += len(batch)
            print(f"  Batch {i // batch_size + 1}: {len(batch)} OK (total: {inserted})")
        except Exception as e:
            error_body = ""
            if hasattr(e, "read"):
                error_body = e.read().decode()
            print(f"  Batch {i // batch_size + 1}: ERROR - {e} {error_body}")
            errors += len(batch)

    print(f"\nDone! Inserted: {inserted}, Errors: {errors}")


def update_missing_fields():
    """Re-parse Excel and update cases that have missing fields."""
    print("Loading Excel for field update...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    # Get all cases with missing description from DB
    url = f"{SUPABASE_URL}/rest/v1/cases?select=id,title,case_no&description=is.null"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    }
    req = Request(url, headers=headers)
    resp = urlopen(req)
    missing = json.loads(resp.read().decode())
    missing_by_title = {c["title"].strip(): c["id"] for c in missing if c.get("title")}
    missing_by_no = {str(c["case_no"]).strip(): c["id"] for c in missing if c.get("case_no")}
    print(f"Cases with missing description: {len(missing)}")

    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        content = row[5]
        if not content:
            continue

        col_title = row[4]
        row_no = row[0]
        title = str(col_title).strip() if col_title else ""
        case_no = str(row_no).strip() if row_no else ""

        # Match by title or case_no
        case_id = None
        if title and title in missing_by_title:
            case_id = missing_by_title[title]
        elif case_no and case_no in missing_by_no:
            case_id = missing_by_no[case_no]
        else:
            continue

        parsed = parse_content(str(content))
        if not parsed.get("description") and not parsed.get("fee") and not parsed.get("must_req"):
            continue

        update_data = {}
        fee = extract_fee(parsed.get("fee", ""))
        if fee:
            update_data["fee"] = fee
        if parsed.get("description"):
            update_data["description"] = parsed["description"][:5000]
        if parsed.get("location"):
            update_data["location"] = parsed["location"][:500]
        if parsed.get("occupancy"):
            update_data["occupancy"] = parsed["occupancy"][:200]
        if parsed.get("must_req"):
            update_data["must_req"] = parsed["must_req"][:5000]
        if parsed.get("nice_to_have"):
            update_data["nice_to_have"] = parsed["nice_to_have"][:5000]
        if parsed.get("flow"):
            update_data["flow"] = parsed["flow"][:200]
        if parsed.get("start_date"):
            update_data["start_date"] = parsed["start_date"][:200]
        if parsed.get("office_days"):
            update_data["office_days"] = parsed["office_days"][:200]

        if not update_data:
            continue

        from urllib.parse import quote
        patch_url = f"{SUPABASE_URL}/rest/v1/cases?id=eq.{case_id}"
        patch_headers = {
            "apikey": SUPABASE_SERVICE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }
        body = json.dumps(update_data, ensure_ascii=False).encode("utf-8")
        req = Request(patch_url, data=body, headers=patch_headers, method="PATCH")
        try:
            urlopen(req)
            updated += 1
        except Exception as e:
            print(f"  Update error for {(title or case_no)[:40]}: {e}")

    print(f"Updated {updated} cases with missing fields.")


if __name__ == "__main__":
    if "--update" in sys.argv:
        update_missing_fields()
    else:
        main()
